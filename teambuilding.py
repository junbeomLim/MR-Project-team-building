from openpyxl import load_workbook

dic = {"자전":"자전거 가속장치(홍길동)","주행":"자율주행차(홍길동)","K":"K-IRONMAN(홍길동)","모션":"모션 카피 휴머노이드(홍길동)","고양":"고양이 관리 세트(홍길동)","경사":"계단/경사 극복 배달로봇(홍길동) / 엘리베이터 타는 배달로봇(홍길동)","애벌":"애벌레/지렁이 생체 모방 로봇(홍길동)","UAV":"자율 비행 UAV(홍길동)","수레":"무중력 수레(홍길동)","포수":"포수로봇(홍길동)","기타":"기타 연주 로봇 JIMMYbot(홍길동)","웨어러블":"재활용 웨어러블 로봇(홍길동)","굴렁":"구르는 로봇 굴렁이(홍길동)","그물":"그물 오르는 보행로봇(홍길동)","세우":"(물병을) 잘 세우는 로봇(홍길동)","외팔":"그림 그리는 로봇 외팔의 화가(홍길동)","보조":"교통 약자를 위한 보행 보조 로봇(홍길동)"}

score = {}
for val in dic.values():
    score[val] = 0
#print(score)

load_wb = load_workbook("정기 프로젝트 팀빌딩(응답).xlsx")

sheet = load_wb['설문지 응답 시트1']

row_max = sheet.max_row
#print(row_max)

nfl = []

#선택이 담기는 col 범위
first_choice = 'D'
last_choice = 'H'

for i in range(ord(first_choice),ord(last_choice)+1):
    for row in range(2,row_max+1):
        a = False
        for k in dic.keys():
            if sheet[chr(i)+str(row)].value.find(k) != -1:
                score[dic[k]] += (ord(last_choice)+1-i)
                a = True
                continue
        if a == False:
            nfl.append([sheet[chr(i)+str(row)].value,ord(last_choice)+1-i])

sorted_dict = sorted(score.items(), key = lambda item: item[1], reverse = True)
#print(sorted_dict)
'''
for i in sorted_dict:
    print(list(i))
'''

print("고려 안된 점수 사항 목록")

nfl_2 = []
for v in nfl:
    print(v[1],'점:',v[0])
    v[0] = input("대체 단어: ")
    for k in score.keys():
        if k.find(v[0]) != -1:
            score[dic[v[0]]] += v[1]
            a = True
            continue
    if a == False:
        nfl_2.append(v[0],v[1])

sorted_dict = sorted(score.items(), key = lambda item: item[1], reverse = True)
#print(sorted_dict)
for i in sorted_dict:
    print(i)
print(nfl_2)
